from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

pathWay = 'execl/'

wb = load_workbook(pathWay + '数据记录.xlsx')

#alway copy then do the work
source = wb.active
target = wb.copy_worksheet(source)

#example of getting the cell value
c = source.cell(row= 2,column = 3)

c.value = datetime.date.today()

print(c.value)

